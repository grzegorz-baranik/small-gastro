"""
Reports Service

Generates report data for:
- Daily summary reports (podsumowanie dnia)
- Monthly trends reports (trendy miesieczne)
- Ingredient usage reports (zuzycie skladnikow)
- Spoilage reports (straty)

Also provides Excel export functionality using openpyxl.
"""

import logging
from io import BytesIO
from datetime import date
from decimal import Decimal
from typing import Optional

from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.models.daily_record import DailyRecord, DayStatus
from app.models.inventory_snapshot import InventorySnapshot, SnapshotType, InventoryLocation
from app.models.ingredient import Ingredient
from app.models.delivery import Delivery
from app.models.storage_transfer import StorageTransfer
from app.models.spoilage import Spoilage, SpoilageReason
from app.models.calculated_sale import CalculatedSale

from app.schemas.reports import (
    DailySummaryReportResponse,
    DailySummaryInventoryItem,
    DailySummaryProductItem,
    DailySummaryFinancials,
    DailySummaryDiscrepancyAlert,
    MonthlyTrendsReportResponse,
    MonthlyTrendItem,
    MonthlyBestWorstDay,
    IngredientUsageReportResponse,
    IngredientUsageReportItem,
    IngredientUsageSummaryItem,
    SpoilageReportResponse,
    SpoilageReportItem,
    SpoilageByReasonSummary,
    SpoilageByIngredientSummary,
)

logger = logging.getLogger(__name__)


# -----------------------------------------------------------------------------
# Polish Day Names and Spoilage Reason Labels
# -----------------------------------------------------------------------------

POLISH_DAY_NAMES = {
    0: "Poniedzialek",
    1: "Wtorek",
    2: "Sroda",
    3: "Czwartek",
    4: "Piatek",
    5: "Sobota",
    6: "Niedziela",
}

SPOILAGE_REASON_LABELS = {
    "expired": "Przeterminowany",
    "over_prepared": "Nadmiernie przygotowany",
    "contaminated": "Zanieczyszczony",
    "equipment_failure": "Awaria sprzetu",
    "other": "Inne",
}

# Excel styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CURRENCY_FORMAT = '#,##0.00 "PLN"'
DECIMAL_FORMAT = "#,##0.000"


def _get_polish_day_name(d: date) -> str:
    """Get Polish name of the day of week."""
    return POLISH_DAY_NAMES.get(d.weekday(), "")


def _get_spoilage_reason_label(reason: str) -> str:
    """Get Polish label for spoilage reason."""
    return SPOILAGE_REASON_LABELS.get(reason, reason)


# -----------------------------------------------------------------------------
# Helper Functions
# -----------------------------------------------------------------------------

def _get_ingredient_day_quantities(
    db: Session,
    daily_record_id: int,
    ingredient_id: int
) -> tuple[Decimal, Decimal, Decimal]:
    """
    Get total deliveries, transfers, and spoilage for an ingredient on a day.
    Returns (deliveries_total, transfers_total, spoilage_total).
    """
    deliveries_sum = db.query(func.coalesce(func.sum(Delivery.quantity), 0)).filter(
        Delivery.daily_record_id == daily_record_id,
        Delivery.ingredient_id == ingredient_id
    ).scalar()

    transfers_sum = db.query(func.coalesce(func.sum(StorageTransfer.quantity), 0)).filter(
        StorageTransfer.daily_record_id == daily_record_id,
        StorageTransfer.ingredient_id == ingredient_id
    ).scalar()

    spoilage_sum = db.query(func.coalesce(func.sum(Spoilage.quantity), 0)).filter(
        Spoilage.daily_record_id == daily_record_id,
        Spoilage.ingredient_id == ingredient_id
    ).scalar()

    return (
        Decimal(str(deliveries_sum)),
        Decimal(str(transfers_sum)),
        Decimal(str(spoilage_sum))
    )


def _calculate_discrepancy_level(discrepancy_percent: Optional[Decimal]) -> Optional[str]:
    """Determine discrepancy level based on percentage threshold."""
    if discrepancy_percent is None:
        return None
    abs_percent = abs(discrepancy_percent)
    if abs_percent <= 5:
        return "ok"
    elif abs_percent <= 10:
        return "warning"
    else:
        return "critical"


def _apply_excel_header_style(ws, row: int, col_count: int):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_adjust_column_widths(ws):
    """Auto-adjust column widths based on content."""
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


# -----------------------------------------------------------------------------
# Daily Summary Report
# -----------------------------------------------------------------------------

def get_daily_summary_report(
    db: Session,
    record_id: int
) -> Optional[DailySummaryReportResponse]:
    """
    Generate a full daily summary report for a given record.

    Returns None if record not found.
    """
    # Get the daily record
    record = db.query(DailyRecord).filter(DailyRecord.id == record_id).first()
    if not record:
        return None

    # Get opening snapshots
    opening_snapshots = db.query(InventorySnapshot).filter(
        InventorySnapshot.daily_record_id == record_id,
        InventorySnapshot.snapshot_type == SnapshotType.OPEN,
        InventorySnapshot.location == InventoryLocation.SHOP
    ).all()

    # Get closing snapshots (if day is closed)
    closing_map: dict[int, Decimal] = {}
    if record.status == DayStatus.CLOSED:
        closing_snapshots = db.query(InventorySnapshot).filter(
            InventorySnapshot.daily_record_id == record_id,
            InventorySnapshot.snapshot_type == SnapshotType.CLOSE,
            InventorySnapshot.location == InventoryLocation.SHOP
        ).all()
        closing_map = {s.ingredient_id: Decimal(str(s.quantity)) for s in closing_snapshots}

    # Build inventory items
    inventory_items: list[DailySummaryInventoryItem] = []
    discrepancy_alerts: list[DailySummaryDiscrepancyAlert] = []

    for opening_snap in opening_snapshots:
        ingredient = opening_snap.ingredient
        ingredient_id = ingredient.id
        opening_qty = Decimal(str(opening_snap.quantity))
        closing_qty = closing_map.get(ingredient_id, Decimal("0"))

        # Get mid-day quantities
        deliveries, transfers, spoilage = _get_ingredient_day_quantities(
            db, record_id, ingredient_id
        )

        # Calculate expected closing and usage
        expected = opening_qty + deliveries + transfers - spoilage
        usage = opening_qty + deliveries + transfers - spoilage - closing_qty

        # Calculate discrepancy
        discrepancy = None
        discrepancy_percent = None
        discrepancy_level = None
        if record.status == DayStatus.CLOSED and expected != Decimal("0"):
            discrepancy = expected - closing_qty
            discrepancy_percent = (discrepancy / expected) * 100
            discrepancy_level = _calculate_discrepancy_level(discrepancy_percent)

            # Add alert if not ok
            if discrepancy_level and discrepancy_level != "ok":
                message = f"Roznica {discrepancy_percent:.1f}% dla {ingredient.name}"
                discrepancy_alerts.append(DailySummaryDiscrepancyAlert(
                    ingredient_id=ingredient_id,
                    ingredient_name=ingredient.name,
                    discrepancy_percent=discrepancy_percent,
                    level=discrepancy_level,
                    message=message,
                ))

        inventory_items.append(DailySummaryInventoryItem(
            ingredient_id=ingredient_id,
            ingredient_name=ingredient.name,
            unit_label=ingredient.unit_label or "szt",
            opening=opening_qty,
            deliveries=deliveries,
            transfers=transfers,
            spoilage=spoilage,
            closing=closing_qty,
            usage=usage,
            discrepancy=discrepancy,
            discrepancy_percent=discrepancy_percent,
            discrepancy_level=discrepancy_level,
        ))

    # Get calculated sales (products sold)
    products_sold: list[DailySummaryProductItem] = []
    if record.status == DayStatus.CLOSED:
        sales = db.query(CalculatedSale).filter(
            CalculatedSale.daily_record_id == record_id
        ).all()

        for sale in sales:
            variant = sale.product_variant
            product = variant.product
            products_sold.append(DailySummaryProductItem(
                product_id=product.id,
                product_name=product.name,
                variant_id=variant.id,
                variant_name=variant.name,
                quantity_sold=Decimal(str(sale.quantity_sold)),
                unit_price_pln=Decimal(str(variant.price_pln)),
                revenue_pln=Decimal(str(sale.revenue_pln)),
            ))

    # Calculate financials
    total_income = Decimal(str(record.total_income_pln or 0))
    total_delivery_cost = Decimal(str(record.total_delivery_cost_pln or 0))
    total_spoilage_cost = Decimal(str(record.total_spoilage_cost_pln or 0))
    net_profit = total_income - total_delivery_cost - total_spoilage_cost

    financials = DailySummaryFinancials(
        total_income_pln=total_income,
        total_delivery_cost_pln=total_delivery_cost,
        total_spoilage_cost_pln=total_spoilage_cost,
        net_profit_pln=net_profit,
    )

    # Format times
    opening_time = record.opened_at.strftime("%H:%M") if record.opened_at else None
    closing_time = record.closed_at.strftime("%H:%M") if record.closed_at else None

    return DailySummaryReportResponse(
        record_id=record.id,
        date=record.date,
        day_of_week=_get_polish_day_name(record.date),
        opening_time=opening_time,
        closing_time=closing_time,
        status=record.status.value,
        inventory_items=inventory_items,
        products_sold=products_sold,
        financials=financials,
        discrepancy_alerts=discrepancy_alerts,
        notes=record.notes,
    )


def export_daily_summary_to_excel(report: DailySummaryReportResponse) -> BytesIO:
    """Export daily summary report to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Podsumowanie dnia"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Podsumowanie dnia - {report.date} ({report.day_of_week})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Basic info
    ws["A3"] = "Godzina otwarcia:"
    ws["B3"] = report.opening_time or "-"
    ws["A4"] = "Godzina zamkniecia:"
    ws["B4"] = report.closing_time or "-"
    ws["A5"] = "Status:"
    ws["B5"] = report.status

    # Inventory table
    row = 7
    ws[f"A{row}"] = "STAN MAGAZYNOWY"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    headers = ["Skladnik", "Jednostka", "Otwarcie", "Dostawy", "Transfery", "Straty", "Zamkniecie", "Zuzycie"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _apply_excel_header_style(ws, row, len(headers))
    row += 1

    for item in report.inventory_items:
        ws.cell(row=row, column=1, value=item.ingredient_name)
        ws.cell(row=row, column=2, value=item.unit_label)
        ws.cell(row=row, column=3, value=float(item.opening))
        ws.cell(row=row, column=4, value=float(item.deliveries))
        ws.cell(row=row, column=5, value=float(item.transfers))
        ws.cell(row=row, column=6, value=float(item.spoilage))
        ws.cell(row=row, column=7, value=float(item.closing))
        ws.cell(row=row, column=8, value=float(item.usage))
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    row += 1

    # Products sold table
    ws[f"A{row}"] = "SPRZEDANE PRODUKTY"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    headers = ["Produkt", "Wariant", "Ilosc", "Cena jedn.", "Przychod"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _apply_excel_header_style(ws, row, len(headers))
    row += 1

    for item in report.products_sold:
        ws.cell(row=row, column=1, value=item.product_name)
        ws.cell(row=row, column=2, value=item.variant_name or "-")
        ws.cell(row=row, column=3, value=float(item.quantity_sold))
        cell = ws.cell(row=row, column=4, value=float(item.unit_price_pln))
        cell.number_format = CURRENCY_FORMAT
        cell = ws.cell(row=row, column=5, value=float(item.revenue_pln))
        cell.number_format = CURRENCY_FORMAT
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    row += 1

    # Financial summary
    ws[f"A{row}"] = "PODSUMOWANIE FINANSOWE"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    ws.cell(row=row, column=1, value="Calkowity przychod:")
    cell = ws.cell(row=row, column=2, value=float(report.financials.total_income_pln))
    cell.number_format = CURRENCY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Koszty dostaw:")
    cell = ws.cell(row=row, column=2, value=float(report.financials.total_delivery_cost_pln))
    cell.number_format = CURRENCY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Koszty strat:")
    cell = ws.cell(row=row, column=2, value=float(report.financials.total_spoilage_cost_pln))
    cell.number_format = CURRENCY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Zysk netto:")
    ws[f"A{row}"].font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value=float(report.financials.net_profit_pln))
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True)
    row += 1

    # Discrepancy alerts
    if report.discrepancy_alerts:
        row += 1
        ws[f"A{row}"] = "ALERTY ROZBIEZNOSCI"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FF0000")
        row += 1

        for alert in report.discrepancy_alerts:
            ws.cell(row=row, column=1, value=alert.message)
            ws[f"A{row}"].font = Font(color="FF0000" if alert.level == "critical" else "FF8C00")
            row += 1

    _auto_adjust_column_widths(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# -----------------------------------------------------------------------------
# Monthly Trends Report
# -----------------------------------------------------------------------------

def get_monthly_trends_report(
    db: Session,
    start_date: date,
    end_date: date
) -> MonthlyTrendsReportResponse:
    """
    Generate monthly trends report for a date range.

    Aggregates daily income, costs, and profit.
    """
    # Get all closed daily records in range
    records = db.query(DailyRecord).filter(
        DailyRecord.date >= start_date,
        DailyRecord.date <= end_date,
        DailyRecord.status == DayStatus.CLOSED
    ).order_by(DailyRecord.date).all()

    items: list[MonthlyTrendItem] = []
    total_income = Decimal("0")
    total_delivery_cost = Decimal("0")
    total_spoilage_cost = Decimal("0")

    best_day: Optional[MonthlyBestWorstDay] = None
    worst_day: Optional[MonthlyBestWorstDay] = None
    best_profit = None
    worst_profit = None

    for record in records:
        income = Decimal(str(record.total_income_pln or 0))
        delivery_cost = Decimal(str(record.total_delivery_cost_pln or 0))
        spoilage_cost = Decimal(str(record.total_spoilage_cost_pln or 0))
        profit = income - delivery_cost - spoilage_cost

        items.append(MonthlyTrendItem(
            date=record.date,
            day_of_week=_get_polish_day_name(record.date),
            income_pln=income,
            delivery_cost_pln=delivery_cost,
            spoilage_cost_pln=spoilage_cost,
            profit_pln=profit,
        ))

        total_income += income
        total_delivery_cost += delivery_cost
        total_spoilage_cost += spoilage_cost

        # Track best/worst days
        if best_profit is None or profit > best_profit:
            best_profit = profit
            best_day = MonthlyBestWorstDay(
                date=record.date,
                day_of_week=_get_polish_day_name(record.date),
                profit_pln=profit,
            )

        if worst_profit is None or profit < worst_profit:
            worst_profit = profit
            worst_day = MonthlyBestWorstDay(
                date=record.date,
                day_of_week=_get_polish_day_name(record.date),
                profit_pln=profit,
            )

    total_profit = total_income - total_delivery_cost - total_spoilage_cost
    days_count = len(items)
    avg_daily_income = total_income / days_count if days_count > 0 else Decimal("0")
    avg_daily_profit = total_profit / days_count if days_count > 0 else Decimal("0")

    return MonthlyTrendsReportResponse(
        start_date=start_date,
        end_date=end_date,
        items=items,
        total_income_pln=total_income,
        total_delivery_cost_pln=total_delivery_cost,
        total_spoilage_cost_pln=total_spoilage_cost,
        total_profit_pln=total_profit,
        avg_daily_income_pln=avg_daily_income,
        avg_daily_profit_pln=avg_daily_profit,
        best_day=best_day,
        worst_day=worst_day,
        days_count=days_count,
    )


def export_monthly_trends_to_excel(report: MonthlyTrendsReportResponse) -> BytesIO:
    """Export monthly trends report to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Trendy miesieczne"

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Trendy miesieczne - {report.start_date} do {report.end_date}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Summary
    row = 3
    ws.cell(row=row, column=1, value="Liczba dni:")
    ws.cell(row=row, column=2, value=report.days_count)
    row += 1
    ws.cell(row=row, column=1, value="Calkowity przychod:")
    cell = ws.cell(row=row, column=2, value=float(report.total_income_pln))
    cell.number_format = CURRENCY_FORMAT
    row += 1
    ws.cell(row=row, column=1, value="Koszty dostaw:")
    cell = ws.cell(row=row, column=2, value=float(report.total_delivery_cost_pln))
    cell.number_format = CURRENCY_FORMAT
    row += 1
    ws.cell(row=row, column=1, value="Koszty strat:")
    cell = ws.cell(row=row, column=2, value=float(report.total_spoilage_cost_pln))
    cell.number_format = CURRENCY_FORMAT
    row += 1
    ws.cell(row=row, column=1, value="Calkowity zysk:")
    ws[f"A{row}"].font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value=float(report.total_profit_pln))
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Sredni dzienny przychod:")
    cell = ws.cell(row=row, column=2, value=float(report.avg_daily_income_pln))
    cell.number_format = CURRENCY_FORMAT
    row += 1
    ws.cell(row=row, column=1, value="Sredni dzienny zysk:")
    cell = ws.cell(row=row, column=2, value=float(report.avg_daily_profit_pln))
    cell.number_format = CURRENCY_FORMAT
    row += 2

    if report.best_day:
        ws.cell(row=row, column=1, value="Najlepszy dzien:")
        ws.cell(row=row, column=2, value=f"{report.best_day.date} ({report.best_day.day_of_week})")
        cell = ws.cell(row=row, column=3, value=float(report.best_day.profit_pln))
        cell.number_format = CURRENCY_FORMAT
        row += 1

    if report.worst_day:
        ws.cell(row=row, column=1, value="Najgorszy dzien:")
        ws.cell(row=row, column=2, value=f"{report.worst_day.date} ({report.worst_day.day_of_week})")
        cell = ws.cell(row=row, column=3, value=float(report.worst_day.profit_pln))
        cell.number_format = CURRENCY_FORMAT
        row += 2

    # Daily data table
    ws[f"A{row}"] = "DANE DZIENNE"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    headers = ["Data", "Dzien", "Przychod", "Koszty dostaw", "Straty", "Zysk"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _apply_excel_header_style(ws, row, len(headers))
    row += 1

    for item in report.items:
        ws.cell(row=row, column=1, value=str(item.date))
        ws.cell(row=row, column=2, value=item.day_of_week)
        cell = ws.cell(row=row, column=3, value=float(item.income_pln))
        cell.number_format = CURRENCY_FORMAT
        cell = ws.cell(row=row, column=4, value=float(item.delivery_cost_pln))
        cell.number_format = CURRENCY_FORMAT
        cell = ws.cell(row=row, column=5, value=float(item.spoilage_cost_pln))
        cell.number_format = CURRENCY_FORMAT
        cell = ws.cell(row=row, column=6, value=float(item.profit_pln))
        cell.number_format = CURRENCY_FORMAT
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    _auto_adjust_column_widths(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# -----------------------------------------------------------------------------
# Ingredient Usage Report
# -----------------------------------------------------------------------------

def get_ingredient_usage_report(
    db: Session,
    start_date: date,
    end_date: date,
    ingredient_ids: Optional[list[int]] = None
) -> IngredientUsageReportResponse:
    """
    Generate ingredient usage report for a date range.

    Optionally filters by specific ingredients.
    """
    # Get all closed daily records in range
    records = db.query(DailyRecord).filter(
        DailyRecord.date >= start_date,
        DailyRecord.date <= end_date,
        DailyRecord.status == DayStatus.CLOSED
    ).order_by(DailyRecord.date).all()

    items: list[IngredientUsageReportItem] = []
    usage_totals: dict[int, dict] = {}  # ingredient_id -> {total_used, days, name, unit_label}

    for record in records:
        # Get opening snapshots for this day
        opening_query = db.query(InventorySnapshot).filter(
            InventorySnapshot.daily_record_id == record.id,
            InventorySnapshot.snapshot_type == SnapshotType.OPEN,
            InventorySnapshot.location == InventoryLocation.SHOP
        )

        # Filter by ingredient IDs if provided
        if ingredient_ids:
            opening_query = opening_query.filter(
                InventorySnapshot.ingredient_id.in_(ingredient_ids)
            )

        opening_snapshots = opening_query.all()

        # Get closing snapshots
        closing_snapshots = db.query(InventorySnapshot).filter(
            InventorySnapshot.daily_record_id == record.id,
            InventorySnapshot.snapshot_type == SnapshotType.CLOSE,
            InventorySnapshot.location == InventoryLocation.SHOP
        ).all()
        closing_map = {s.ingredient_id: Decimal(str(s.quantity)) for s in closing_snapshots}

        for opening_snap in opening_snapshots:
            ingredient = opening_snap.ingredient
            ingredient_id = ingredient.id

            # Skip if filtering and not in list
            if ingredient_ids and ingredient_id not in ingredient_ids:
                continue

            opening_qty = Decimal(str(opening_snap.quantity))
            closing_qty = closing_map.get(ingredient_id, Decimal("0"))

            # Get mid-day quantities
            deliveries, transfers, spoilage = _get_ingredient_day_quantities(
                db, record.id, ingredient_id
            )

            # Calculate usage
            expected = opening_qty + deliveries + transfers - spoilage
            usage = opening_qty + deliveries + transfers - spoilage - closing_qty

            # Calculate discrepancy
            discrepancy = None
            discrepancy_percent = None
            if expected != Decimal("0"):
                discrepancy = expected - closing_qty
                discrepancy_percent = (discrepancy / expected) * 100

            items.append(IngredientUsageReportItem(
                date=record.date,
                day_of_week=_get_polish_day_name(record.date),
                ingredient_id=ingredient_id,
                ingredient_name=ingredient.name,
                unit_label=ingredient.unit_label or "szt",
                opening=opening_qty,
                deliveries=deliveries,
                transfers=transfers,
                spoilage=spoilage,
                closing=closing_qty,
                usage=usage,
                discrepancy=discrepancy,
                discrepancy_percent=discrepancy_percent,
            ))

            # Accumulate for summary
            if ingredient_id not in usage_totals:
                usage_totals[ingredient_id] = {
                    "name": ingredient.name,
                    "unit_label": ingredient.unit_label or "szt",
                    "total_used": Decimal("0"),
                    "days": 0,
                }
            usage_totals[ingredient_id]["total_used"] += usage
            usage_totals[ingredient_id]["days"] += 1

    # Build summary
    summary: list[IngredientUsageSummaryItem] = []
    for ing_id, data in usage_totals.items():
        avg_daily = data["total_used"] / data["days"] if data["days"] > 0 else Decimal("0")
        summary.append(IngredientUsageSummaryItem(
            ingredient_id=ing_id,
            ingredient_name=data["name"],
            unit_label=data["unit_label"],
            total_used=data["total_used"],
            avg_daily_usage=avg_daily,
            days_with_data=data["days"],
        ))

    return IngredientUsageReportResponse(
        start_date=start_date,
        end_date=end_date,
        filtered_ingredient_ids=ingredient_ids,
        items=items,
        summary=summary,
    )


def export_ingredient_usage_to_excel(report: IngredientUsageReportResponse) -> BytesIO:
    """Export ingredient usage report to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Zuzycie skladnikow"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Zuzycie skladnikow - {report.start_date} do {report.end_date}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Summary section
    row = 3
    ws[f"A{row}"] = "PODSUMOWANIE"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    headers = ["Skladnik", "Jednostka", "Laczne zuzycie", "Srednie dzienne", "Dni z danymi"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _apply_excel_header_style(ws, row, len(headers))
    row += 1

    for item in report.summary:
        ws.cell(row=row, column=1, value=item.ingredient_name)
        ws.cell(row=row, column=2, value=item.unit_label)
        ws.cell(row=row, column=3, value=float(item.total_used))
        ws.cell(row=row, column=4, value=float(item.avg_daily_usage))
        ws.cell(row=row, column=5, value=item.days_with_data)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    row += 1

    # Detailed data
    ws[f"A{row}"] = "SZCZEGOLY DZIENNE"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    headers = ["Data", "Dzien", "Skladnik", "Jedn.", "Otwarcie", "Dostawy", "Transfery", "Straty", "Zamkniecie", "Zuzycie"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _apply_excel_header_style(ws, row, len(headers))
    row += 1

    for item in report.items:
        ws.cell(row=row, column=1, value=str(item.date))
        ws.cell(row=row, column=2, value=item.day_of_week)
        ws.cell(row=row, column=3, value=item.ingredient_name)
        ws.cell(row=row, column=4, value=item.unit_label)
        ws.cell(row=row, column=5, value=float(item.opening))
        ws.cell(row=row, column=6, value=float(item.deliveries))
        ws.cell(row=row, column=7, value=float(item.transfers))
        ws.cell(row=row, column=8, value=float(item.spoilage))
        ws.cell(row=row, column=9, value=float(item.closing))
        ws.cell(row=row, column=10, value=float(item.usage))
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    _auto_adjust_column_widths(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# -----------------------------------------------------------------------------
# Spoilage Report
# -----------------------------------------------------------------------------

def get_spoilage_report(
    db: Session,
    start_date: date,
    end_date: date,
    group_by: str = "date"
) -> SpoilageReportResponse:
    """
    Generate spoilage report for a date range.

    group_by can be 'date', 'ingredient', or 'reason'.
    """
    # Get all spoilage records in range via daily records
    spoilages = (
        db.query(Spoilage)
        .join(DailyRecord)
        .filter(
            DailyRecord.date >= start_date,
            DailyRecord.date <= end_date
        )
        .order_by(DailyRecord.date)
        .all()
    )

    items: list[SpoilageReportItem] = []
    by_reason: dict[str, dict] = {}
    by_ingredient: dict[int, dict] = {}

    for spoilage in spoilages:
        record = spoilage.daily_record
        ingredient = spoilage.ingredient
        reason_value = spoilage.reason.value
        reason_label = _get_spoilage_reason_label(reason_value)

        items.append(SpoilageReportItem(
            id=spoilage.id,
            date=record.date,
            day_of_week=_get_polish_day_name(record.date),
            ingredient_id=ingredient.id,
            ingredient_name=ingredient.name,
            unit_label=ingredient.unit_label or "szt",
            quantity=Decimal(str(spoilage.quantity)),
            reason=reason_value,
            reason_label=reason_label,
            notes=spoilage.notes,
        ))

        # Accumulate by reason
        if reason_value not in by_reason:
            by_reason[reason_value] = {
                "label": reason_label,
                "count": 0,
                "quantity": Decimal("0"),
            }
        by_reason[reason_value]["count"] += 1
        by_reason[reason_value]["quantity"] += Decimal(str(spoilage.quantity))

        # Accumulate by ingredient
        if ingredient.id not in by_ingredient:
            by_ingredient[ingredient.id] = {
                "name": ingredient.name,
                "unit_label": ingredient.unit_label or "szt",
                "count": 0,
                "quantity": Decimal("0"),
            }
        by_ingredient[ingredient.id]["count"] += 1
        by_ingredient[ingredient.id]["quantity"] += Decimal(str(spoilage.quantity))

    # Sort items based on group_by
    if group_by == "ingredient":
        items.sort(key=lambda x: (x.ingredient_name, x.date))
    elif group_by == "reason":
        items.sort(key=lambda x: (x.reason, x.date))
    # else: already sorted by date

    # Build summary lists
    reason_summary = [
        SpoilageByReasonSummary(
            reason=reason,
            reason_label=data["label"],
            total_count=data["count"],
            total_quantity=data["quantity"],
        )
        for reason, data in by_reason.items()
    ]

    ingredient_summary = [
        SpoilageByIngredientSummary(
            ingredient_id=ing_id,
            ingredient_name=data["name"],
            unit_label=data["unit_label"],
            total_count=data["count"],
            total_quantity=data["quantity"],
        )
        for ing_id, data in by_ingredient.items()
    ]

    return SpoilageReportResponse(
        start_date=start_date,
        end_date=end_date,
        group_by=group_by,
        items=items,
        by_reason=reason_summary,
        by_ingredient=ingredient_summary,
        total_count=len(items),
    )


def export_spoilage_to_excel(report: SpoilageReportResponse) -> BytesIO:
    """Export spoilage report to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Straty"

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Raport strat - {report.start_date} do {report.end_date}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Summary by reason
    row = 3
    ws[f"A{row}"] = "PODSUMOWANIE WG PRZYCZYNY"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    headers = ["Przyczyna", "Liczba", "Laczna ilosc"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _apply_excel_header_style(ws, row, len(headers))
    row += 1

    for item in report.by_reason:
        ws.cell(row=row, column=1, value=item.reason_label)
        ws.cell(row=row, column=2, value=item.total_count)
        ws.cell(row=row, column=3, value=float(item.total_quantity))
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    row += 1

    # Summary by ingredient
    ws[f"A{row}"] = "PODSUMOWANIE WG SKLADNIKA"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    headers = ["Skladnik", "Jednostka", "Liczba", "Laczna ilosc"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _apply_excel_header_style(ws, row, len(headers))
    row += 1

    for item in report.by_ingredient:
        ws.cell(row=row, column=1, value=item.ingredient_name)
        ws.cell(row=row, column=2, value=item.unit_label)
        ws.cell(row=row, column=3, value=item.total_count)
        ws.cell(row=row, column=4, value=float(item.total_quantity))
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    row += 1

    # Detailed data
    ws[f"A{row}"] = "SZCZEGOLY"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    headers = ["Data", "Dzien", "Skladnik", "Jedn.", "Ilosc", "Przyczyna", "Notatki"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _apply_excel_header_style(ws, row, len(headers))
    row += 1

    for item in report.items:
        ws.cell(row=row, column=1, value=str(item.date))
        ws.cell(row=row, column=2, value=item.day_of_week)
        ws.cell(row=row, column=3, value=item.ingredient_name)
        ws.cell(row=row, column=4, value=item.unit_label)
        ws.cell(row=row, column=5, value=float(item.quantity))
        ws.cell(row=row, column=6, value=item.reason_label)
        ws.cell(row=row, column=7, value=item.notes or "")
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    _auto_adjust_column_widths(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
